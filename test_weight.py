import pyvisa
import time
import pandas as pd
from datetime import datetime
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# ================= CONFIG =================

INTERVAL_SEC = 0.5
PLOT_UPDATE_INTERVAL = 5
PRINT_INTERVAL_SEC = 3.0

OUTPUT_DIR = "battery_test_results"
FILENAME = f"{datetime.now().strftime('%y%m%d_%H%M')}_test_DL3021"

DISCHARGE_CURRENT_A = 1.5
CUT_OFF_VOLTAGE = 3.4

SENSE = True

pd.set_option('display.float_format', '{:.2f}'.format)

# ================= OPTIONAL WEIGHT =================

weight_input = input("Battery weight in grams? (press Enter to skip): ").strip()
BATTERY_WEIGHT_G = float(weight_input) if weight_input else None

# ================= SAVE DIR =================

save_dir = os.path.join(OUTPUT_DIR, FILENAME)
os.makedirs(save_dir, exist_ok=True)

# ================= VISA =================

rm = pyvisa.ResourceManager()

rigol_resource = None
for res in rm.list_resources():
    if "DL3" in res or "0x1AB1" in res:
        rigol_resource = res
        break

if not rigol_resource:
    print("❌ DL3021 not found")
    exit()

inst = rm.open_resource(rigol_resource)

inst.timeout = 200
inst.read_termination = '\n'
inst.write_termination = '\n'

print("Connected:", inst.query("*IDN?").strip())

# ================= SETUP =================

inst.write("*RST")
inst.query("*OPC?")

inst.write(":SENS ON" if SENSE else ":SENS OFF")

inst.write(":FUNC CC")
inst.write(f":CURR {DISCHARGE_CURRENT_A}")
inst.write(":INPUT ON")

print("🚀 Start CC Discharge")

# ================= DATA =================

data = []

cap_ah = 0.0
energy_wh = 0.0

prev = time.time()
elapsed_sec = 0.0

# ================= PLOT =================

plt.ion()
fig, ax = plt.subplots()

line, = ax.plot([], [], lw=2)

ax.set_ylim(2.5, 4.3)
ax.set_xlabel("Time (min)")
ax.set_ylabel("Voltage (V)")

x, y = [], []

# ================= LOOP =================

i = 0
last_print = time.time()

try:
    while True:

        loop_start = time.time()

        dt_sec = loop_start - prev
        dt_hr = dt_sec / 3600.0
        prev = loop_start
        elapsed_sec += dt_sec

        t = datetime.now()

        # ===== MEASUREMENT =====
        inst.write(":MEAS:VOLT?")
        voltage = float(inst.read())

        inst.write(":MEAS:CURR?")
        current = float(inst.read())

        # ===== POWER =====
        power = voltage * current

        # ===== RESISTANCE =====
        resistance = voltage / current if abs(current) > 1e-6 else None

        # ===== CAPACITY / ENERGY =====
        cap_ah += current * dt_hr
        energy_wh += power * dt_hr
        cap_mAh = cap_ah * 1000.0

        # ===== STORE RAW (SEC 유지) =====
        data.append([
            elapsed_sec,   # 🔥 RAW seconds 그대로 저장
            voltage,
            current,
            power,
            resistance,
            cap_mAh,
            energy_wh
        ])

        # ===== PLOT (MINUTE) =====
        if i % PLOT_UPDATE_INTERVAL == 0:
            x.append(elapsed_sec / 60.0)   # 🔥 min 변환
            y.append(voltage)

            line.set_data(x, y)

            ax.set_xlim(
                max(0, (elapsed_sec - 30) / 60.0),
                (elapsed_sec / 60.0) + 0.1
            )

            plt.pause(0.0001)

        # ===== PRINT =====
        if time.time() - last_print >= PRINT_INTERVAL_SEC:
            res_str = f"{resistance:.2f} Ω" if resistance is not None else "NaN Ω"
            print(
                f"{t.strftime('%H:%M:%S')} | "
                f"V={voltage:.2f} V | "
                f"I={current:.2f} A | "
                f"P={power:.2f} W | "
                f"R={res_str} | "
                f"Cap={cap_mAh:.2f} mAh | "
                f"E={energy_wh:.2f} Wh"
            )
            last_print = time.time()

        # ===== CUT OFF =====
        if voltage <= CUT_OFF_VOLTAGE:
            print("\n✅ Cutoff reached")
            break

        i += 1

        remain = INTERVAL_SEC - (time.time() - loop_start)
        if remain > 0:
            time.sleep(remain)

except KeyboardInterrupt:
    print("\n🛑 Stopped by user")

finally:

    try:
        inst.write(":INPUT OFF")
        inst.close()
    except:
        pass

    if len(data) == 0:
        print("❌ No data collected")
        exit()

    # ================= DATAFRAME =================

    df = pd.DataFrame(
        data,
        columns=[
            "Time(sec)",   # 🔥 RAW sec
            "Voltage(V)",
            "Current(A)",
            "Power(W)",
            "Resistance(Ohm)",
            "Capacity(mAh)",
            "Energy(Wh)"
        ]
    )

    # ===== CLEAN =====
    df["Resistance(Ohm)"] = pd.to_numeric(df["Resistance(Ohm)"], errors="coerce")

    # ===== ROUNDING =====
    cols = [
        "Time(sec)",
        "Voltage(V)",
        "Current(A)",
        "Power(W)",
        "Resistance(Ohm)",
        "Capacity(mAh)",
        "Energy(Wh)"
    ]

    df[cols] = df[cols].round(2)

    # ================= SAVE RAW =================

    raw_path = os.path.join(save_dir, "raw.xlsx")
    summary_path = os.path.join(save_dir, "summary.xlsx")

    df.to_excel(raw_path, index=False)

    wb = load_workbook(raw_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    wb.save(raw_path)

    # ================= AVG POWER =================

    total_time_hr = df["Time(sec)"].iloc[-1] / 3600.0
    avg_power = df["Energy(Wh)"].iloc[-1] / total_time_hr

    # ================= SUMMARY =================

    summary_dict = {

        "AVG V": round(df["Voltage(V)"].mean(), 2),
        "I (A)": round(df["Current(A)"].mean(), 2),
        "Avg Power(W)": round(avg_power, 2),

        "Max V": round(df["Voltage(V)"].max(), 2),
        "Final V": round(df["Voltage(V)"].iloc[-1], 2),

        "Capacity(mAh)": round(df["Capacity(mAh)"].iloc[-1], 2),
        "Energy(Wh)": round(df["Energy(Wh)"].iloc[-1], 2),

        "Avg Resistance(Ohm)": round(df["Resistance(Ohm)"].mean(skipna=True), 2),
        # "Max Resistance(Ohm)": round(df["Resistance(Ohm)"].max(skipna=True), 2),
        # "Min Resistance(Ohm)": round(df["Resistance(Ohm)"].min(skipna=True), 2),

        "Total Time(min)": round(df["Time(sec)"].iloc[-1] / 60.0, 2),
    }

    if BATTERY_WEIGHT_G:

        summary_dict["Battery Weight(g)"] = BATTERY_WEIGHT_G
        summary_dict["Energy Density(Wh/kg)"] = round(
            (df["Energy(Wh)"].iloc[-1] * 1000) / BATTERY_WEIGHT_G, 2
        )
        summary_dict["Power Density(W/kg)"] = round(
            (avg_power * 1000) / BATTERY_WEIGHT_G, 2
        )

    summary = pd.DataFrame([summary_dict])
    summary.to_excel(summary_path, index=False)

    # ================= FINAL PLOT =================

    ax.set_xlim(0, (elapsed_sec / 60.0) + 0.1)
    line.set_data(x, y)

    fig.canvas.draw()
    fig.savefig(os.path.join(save_dir, "plot.png"), dpi=200)

    plt.ioff()
    plt.show()

    # ================= PRINT =================

    print("\n================ SUMMARY ================")
    print(summary.to_string(index=False))
    print("=========================================")

    print("\n💾 Saved:")
    print(raw_path)
    print(summary_path)